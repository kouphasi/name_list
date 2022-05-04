from platform import python_branch
import openpyxl as opxl

def devide_3():
    wb = opxl.load_workbook(r"goal\2022_ESS名簿.xlsx")
    ws_a = wb["Member"]
    ws_dra = wb.create_sheet(title="Drama")
    ws_spe = wb.create_sheet(title="Speech")
    ws_dis = wb.create_sheet(title="Discussion")
    wss = [ws_dra, ws_spe, ws_dis]
    for ws in wss:
        ws["A1"] = "学年"
        ws["B1"] = "名前"
        ws["C1"] = "メインorサブ"
    def writeSectionMember(mos,cod):#Main or Sub // C or D(collumn)
        for row in range(2,ws_a.max_row+1):
            sec = ws_a[f"{cod}{row}"].value
            i = 0
            if sec == "ドラマ":
                i = 0
            elif sec == "スピーチ":
                i = 1
            elif sec == "ディスカッション":
                i = 2
            else:
                continue
            
            nextrow = wss[i].max_row + 1
            
            wss[i][f"A{nextrow}"].value = ws_a[f"A{row}"].value
            wss[i][f"B{nextrow}"].value = ws_a[f"B{row}"].value
            wss[i][f"C{nextrow}"].value = mos

    writeSectionMember("メイン","C")
    writeSectionMember("サブ","D")

    wb.save(r"goal\2022_ESS名簿.xlsx")