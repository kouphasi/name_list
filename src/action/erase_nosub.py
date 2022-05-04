import openpyxl as opxl

def erase_noSub():
    wb = opxl.load_workbook(r"goal\2022_ESS名簿.xlsx")
    ws = wb["all member"]

    for row in range(2,ws.max_row+1):
        if ws[f"D{row}"].value == "サブセクションは選択しない":
            ws[f"D{row}"].value = ""

    wb.save(r"goal\2022_ESS名簿.xlsx")