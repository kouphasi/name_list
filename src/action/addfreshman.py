import openpyxl as opxl

def addFresh():
    import datetime
    today = datetime.date.today()

    wb = opxl.load_workbook(r"goal\2022_ESS名簿.xlsx")
    f_wb = opxl.load_workbook(r"sample\2022 新歓 入部届（回答）.xlsx")

    ws = wb.create_sheet("all member")
    ws["A1"].value = "学年"
    ws["B1"].value = "名前"
    ws["C1"].value = "メイン"
    ws["D1"].value = "サブ"
    f_ws = f_wb["フォームの回答 1"]

    for row in range(2,f_ws.max_row+1):
        nextRow = ws.max_row + 1


        ws[f"B{nextRow}"].value = f_ws[f"B{row}"].value
        ws[f"C{nextRow}"].value = f_ws[f"M{row}"].value
        ws[f"D{nextRow}"].value = f_ws[f"N{row}"].value
        
        thisYear = int(today.year)
        Bottom = thisYear - 2000
        
        grade =  Bottom - int(f_ws[f"F{row}"].value[1:3]) + 1 #学籍番号から学年を抽出
        if grade > 1:
            grade = float(grade)-0.5
        ws[f"A{nextRow}"] = grade

    wb.save(r"goal\2022_ESS名簿.xlsx")