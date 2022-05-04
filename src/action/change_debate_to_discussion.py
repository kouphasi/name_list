import openpyxl as opxl

def change_DebToDis():
    wb = opxl.load_workbook(r"goal\2022_ESS名簿.xlsx")
    ws = wb["all member"]

    for row in range(2,ws.max_row+1):
        if ws[f"C{row}"].value == "ディベート":
            ws[f"C{row}"].value = "ディスカッション"
        if ws[f"D{row}"].value == "ディベート":
            if ws[f"C{row}"].value == "ディスカッション":
                ws[f"D{row}"].value == ""
            else:
                ws[f"D{row}"].value = "ディスカッション"

    wb.save(r"goal\2022_ESS名簿.xlsx")