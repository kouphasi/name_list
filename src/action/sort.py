import openpyxl as opxl
from pprint import pprint
from operator import itemgetter

def make_membersheet():
    wb = opxl.load_workbook(r"goal\2022_ESS名簿.xlsx")
    ws = wb["all member"]

    member_list = []
    for row in ws.iter_rows():
        if row[0].row == 1:
            header_cell = row
            
        else:
            row_dict = {}
            for k, v in zip(header_cell, row):
                row_dict[k.value] = v.value
            member_list.append(row_dict)
            

    # print(member_list)
    # pprint(member_list, sort_dicts=False)

    sorted_list_a = sorted(member_list, key=itemgetter("学年"))

    # pprint(sorted_list_a, sort_dicts=False)

    ws_all = wb.create_sheet(title="Member")

    ws_all["A1"].value = "学年"
    ws_all["B1"].value = "名前"
    ws_all["C1"].value = "メイン"
    ws_all["D1"].value = "サブ"

    for i in range(len(sorted_list_a)):
        ws_all[f"A{i+2}"] = sorted_list_a[i]["学年"]
        ws_all[f"B{i+2}"] = sorted_list_a[i]["名前"]
        ws_all[f"C{i+2}"] = sorted_list_a[i]["メイン"]
        ws_all[f"D{i+2}"] = sorted_list_a[i]["サブ"]

    ws.sheet_state = ws.SHEETSTATE_HIDDEN
    wb.remove(wb["Sheet"])
    # ws_all.auto_filter = f"A1:D{ws_all.max_row}"

    wb.save(r"goal\2022_ESS名簿.xlsx")