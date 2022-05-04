import openpyxl as opxl

def add_under4():
    p_wb = opxl.load_workbook(r"sample\2021_ESS名簿.xlsx")
    n_wb = opxl.load_workbook(r"goal\2022_ESS名簿.xlsx")

    p_ws = p_wb["all member"]
    n_ws = n_wb["all member"]

    for row in range(2,p_ws.max_row+1):
        if p_ws[f'A{row}'].value < 4:
            nextRow = n_ws.max_row+1
            n_ws[f"A{nextRow}"] = p_ws[f"A{row}"].value + 1
            n_ws[f"B{nextRow}"].value = p_ws[f"B{row}"].value
            n_ws[f"C{nextRow}"].value = p_ws[f"C{row}"].value
            n_ws[f"D{nextRow}"].value = p_ws[f"D{row}"].value

    n_wb.save(r"goal\2022_ESS名簿.xlsx")