import openpyxl as xl

def doAuto():
    from openpyxl.utils import get_column_letter
    # set input file name
    inputfile = r"goal\2022_ESS名簿.xlsx"

    # read input xlsx
    wb = xl.load_workbook(filename=inputfile)
    for ws in wb:
        # set column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column)].width = str(adjusted_width)

    # save xlsx file
    wb.save(inputfile)